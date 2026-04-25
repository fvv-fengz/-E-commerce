# -*- coding: utf-8 -*-
"""
将 run_template_trial 导出的长表（店铺名、键、标签、数据值）合并为
doc/抖音与拼多多运营看板数据模板.xlsx 同结构的宽表（每店铺一行）。
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

# 采集键 -> 模板列名（与模板表头一致）
KEY_TO_KANBAN_COL: Dict[str, str] = {
    "home_user_pay_amount": "罗盘支付金额",
    "home_deal_order_count": "罗盘成交订单数",
    "home_avg_order_value": "客单价",
    "stat_cost_for_roi2": "千川消耗",
    "total_order_settle_count_for_roi2_1h": "千川净成交订单数",
    "total_prepay_and_pay_settle_roi2_1h": "净成交roi",
}


def template_path(project_root: Optional[Path] = None) -> Path:
    root = project_root or Path(__file__).resolve().parent.parent
    return root / "doc" / "抖音与拼多多运营看板数据模板.xlsx"


def read_template_columns(xlsx_path: Path) -> List[str]:
    """读取模板首行表头，作为输出列顺序。"""
    from openpyxl import load_workbook

    if not xlsx_path.is_file():
        # 兜底（与当前 doc 模板一致）
        return [
            "抖音",
            "店铺",
            "罗盘支付金额",
            "罗盘成交订单数",
            "客单价",
            "千川消耗",
            "千川净成交订单数",
            "净成交roi",
        ]
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers: List[str] = []
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip() != "":
            headers.append(str(v).strip())
    wb.close()
    return headers if headers else list(KEY_TO_KANBAN_COL.values())


def _norm_val(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if s.lower() == "none":
        return ""
    return s


def long_format_to_kanban(
    df_long: pd.DataFrame,
    *,
    template_columns: Optional[List[str]] = None,
) -> pd.DataFrame:
    """
    长表 -> 运营看板宽表。每个店铺一行；同一店铺同一键取最后一次出现的值。
    """
    required = {"店铺名", "键", "数据值"}
    if not required.issubset(set(df_long.columns.astype(str))):
        raise ValueError(f"长表需包含列: {required}")

    cols = template_columns or read_template_columns(template_path())
    # 确保「店铺」在模板中
    shop_col = "店铺"
    if shop_col not in cols:
        cols = ["抖音", "店铺"] + [c for c in cols if c not in ("抖音", "店铺")]

    # 店铺 -> {采集键: 值}
    per_shop: Dict[str, Dict[str, str]] = {}
    for _, row in df_long.iterrows():
        shop = str(row.get("店铺名") or "").strip()
        if not shop:
            continue
        key = str(row.get("键") or "").strip()
        if key not in KEY_TO_KANBAN_COL:
            continue
        val = _norm_val(row.get("数据值"))
        if shop not in per_shop:
            per_shop[shop] = {}
        per_shop[shop][key] = val  # 后者覆盖前者

    all_shops = sorted(
        {str(x).strip() for x in df_long["店铺名"].dropna().tolist() if str(x).strip()}
    )
    out_rows: List[Dict[str, str]] = []
    for shop in all_shops:
        metrics = per_shop.get(shop, {})
        rec = {c: "" for c in cols}
        if "抖音" in rec:
            rec["抖音"] = ""
        if shop_col in rec:
            rec[shop_col] = shop
        for skey, tcol in KEY_TO_KANBAN_COL.items():
            if tcol in rec and skey in metrics:
                rec[tcol] = metrics[skey]
        out_rows.append(rec)

    if not out_rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(out_rows, columns=cols)


def merge_pipeline(df_long: pd.DataFrame) -> pd.DataFrame:
    """读模板列名后转宽表。"""
    tp = template_path()
    headers = read_template_columns(tp)
    return long_format_to_kanban(df_long, template_columns=headers)
